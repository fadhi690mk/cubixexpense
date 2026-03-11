from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth.views import LoginView, PasswordResetView, PasswordResetConfirmView, PasswordChangeView
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView, TemplateView, View
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse_lazy
from django.db.models import Sum, Count, Q, F, DecimalField, Max
from django.db.models.functions import Coalesce
from django.utils import timezone
from datetime import timedelta
from django.http import HttpResponse
from django.contrib.messages.views import SuccessMessageMixin
from django.contrib import messages
import csv
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from .models import Fund, Expense, ExpenseCategory, ExpenseDocument, Contributor, ActivityLog, Todo
from django.contrib.auth import get_user_model

User = get_user_model()


# ==========================================
# AUTHENTICATION VIEWS
# ==========================================

class CustomLoginView(LoginView):
    """Custom login view with Bootstrap styling"""
    template_name = 'auth/login.html'
    redirect_authenticated_user = True
    
    def get_success_url(self):
        return reverse_lazy('dashboard')


class CustomPasswordResetView(PasswordResetView):
    """Custom password reset view"""
    template_name = 'auth/password_reset.html'
    email_template_name = 'auth/password_reset_email.html'
    subject_template_name = 'auth/password_reset_subject.txt'
    success_url = reverse_lazy('login')
    
    def post(self, request, *args, **kwargs):
        response = super().post(request, *args, **kwargs)
        messages.success(request, 'Check your email for password reset instructions.')
        return response


class CustomLogoutView(View):
    """Custom logout view that handles both GET and POST"""
    def get(self, request):
        from django.contrib.auth import logout
        logout(request)
        return redirect('login')

    def post(self, request):
        from django.contrib.auth import logout
        logout(request)
        return redirect('login')


class CustomPasswordChangeView(LoginRequiredMixin, PasswordChangeView):
    """Custom password change view for authenticated users"""
    template_name = 'auth/password_change.html'
    success_url = reverse_lazy('dashboard')
    
    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, 'Your password has been changed successfully!')
        return response


# ==========================================
# DASHBOARD VIEW
# ==========================================

@login_required(login_url='login')
def dashboard(request):
    """Main dashboard showing financial overview"""
    
    # All-time totals
    total_funds = Fund.objects.filter(is_active=True).aggregate(
        total=Coalesce(Sum('amount'), 0, output_field=DecimalField())
    )['total']
    
    total_expenses = Expense.objects.filter(is_active=True).aggregate(
        total=Coalesce(Sum('amount'), 0, output_field=DecimalField())
    )['total']
    
    current_balance = total_funds - total_expenses
    
    # This month's data
    today = timezone.now().date()
    month_start = today.replace(day=1)
    
    month_funds = Fund.objects.filter(
        is_active=True,
        date__gte=month_start
    ).aggregate(
        total=Coalesce(Sum('amount'), 0, output_field=DecimalField())
    )['total']
    
    month_expenses = Expense.objects.filter(
        is_active=True,
        date__gte=month_start
    ).aggregate(
        total=Coalesce(Sum('amount'), 0, output_field=DecimalField())
    )['total']
    
    month_remaining = month_funds - month_expenses
    
    # Recent data
    recent_funds = Fund.objects.filter(is_active=True).order_by('-date')[:5]
    recent_expenses = Expense.objects.filter(is_active=True).order_by('-date')[:5]
    
    # Categories
    active_categories_count = ExpenseCategory.objects.filter(is_active=True).count()
    
    # Todo (Planned Expenses) statistics
    pending_todos_count = Todo.objects.filter(
        is_active=True,
        status='PENDING'
    ).count()
    
    estimated_todos_amount = Todo.objects.filter(
        is_active=True,
        status='PENDING',
        estimated_amount__isnull=False
    ).aggregate(
        total=Coalesce(Sum('estimated_amount'), 0, output_field=DecimalField())
    )['total']
    
    overdue_todos_count = Todo.objects.filter(
        is_active=True,
        status='PENDING',
        target_date__lt=today
    ).count()
    
    completed_todos_count = Todo.objects.filter(
        is_active=True,
        status='COMPLETED',
        updated_at__gte=month_start
    ).count()
    
    # Contributor statistics
    contributor_stats = []
    for contributor in Contributor.objects.filter(is_active=True):
        fund_data = Fund.objects.filter(
            contributor=contributor,
            is_active=True
        ).aggregate(
            total=Coalesce(Sum('amount'), 0, output_field=DecimalField()),
            count=Count('id'),
            last_date=Max('date')
        )
        
        if fund_data['count'] > 0:
            contributor_stats.append({
                'contributor_name': contributor.name,
                'total': fund_data['total'],
                'count': fund_data['count'],
                'last_date': fund_data['last_date'],
            })
    
    context = {
        'total_funds': total_funds,
        'total_expenses': total_expenses,
        'current_balance': current_balance,
        'month_funds': month_funds,
        'month_expenses': month_expenses,
        'month_remaining': month_remaining,
        'active_categories_count': active_categories_count,
        'recent_funds': recent_funds,
        'recent_expenses': recent_expenses,
        'contributor_stats': contributor_stats,
        # Todo stats
        'pending_todos_count': pending_todos_count,
        'estimated_todos_amount': estimated_todos_amount,
        'overdue_todos_count': overdue_todos_count,
        'completed_todos_count': completed_todos_count,
    }
    
    return render(request, 'dashboard.html', context)


# ==========================================
# FUND VIEWS
# ==========================================

class FundListView(LoginRequiredMixin, ListView):
    """List all funds with filtering and pagination"""
    model = Fund
    template_name = 'funds/fund_list.html'
    context_object_name = 'funds'
    paginate_by = 20
    login_url = 'login'
    
    def get_queryset(self):
        queryset = Fund.objects.filter(is_active=True).order_by('-date')
        
        # Filter by contributor
        contributor = self.request.GET.get('contributor')
        if contributor:
            queryset = queryset.filter(contributor_id=contributor)
        
        # Filter by date range
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        
        if date_from:
            queryset = queryset.filter(date__gte=date_from)
        if date_to:
            queryset = queryset.filter(date__lte=date_to)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Calculate stats
        funds = self.get_queryset()
        context['total_amount'] = funds.aggregate(
            total=Coalesce(Sum('amount'), 0, output_field=DecimalField())
        )['total']
        context['funds_count'] = funds.count()
        
        # Get all active contributors for filter dropdown
        context['all_contributors'] = Contributor.objects.filter(is_active=True).order_by('name')
        
        # Calculate contributor totals
        contributor_totals = []
        for contributor in Contributor.objects.filter(is_active=True):
            contributor_total = funds.filter(contributor=contributor).aggregate(
                total=Coalesce(Sum('amount'), 0, output_field=DecimalField())
            )['total']
            if contributor_total > 0:
                contributor_totals.append({
                    'name': contributor.name,
                    'id': contributor.id,
                    'total': contributor_total,
                    'count': funds.filter(contributor=contributor).count()
                })
        
        context['contributor_totals'] = contributor_totals
        
        return context


class FundDetailView(LoginRequiredMixin, DetailView):
    """View fund details"""
    model = Fund
    template_name = 'funds/fund_detail.html'
    context_object_name = 'fund'
    login_url = 'login'
    
    def get_queryset(self):
        return Fund.objects.filter(is_active=True)


class FundCreateView(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    """Create a new fund"""
    model = Fund
    template_name = 'funds/fund_form.html'
    fields = ['contributor', 'amount', 'date', 'description']
    success_url = reverse_lazy('fund-list')
    success_message = 'Fund added successfully!'
    login_url = 'login'
    
    def form_valid(self, form):
        form.instance.added_by = self.request.user
        response = super().form_valid(form)
        
        # Log activity
        ActivityLog.objects.create(
            user=self.request.user,
            action='CREATE',
            model_name='Fund',
            object_id=self.object.id,
            description=f'Created fund of {self.object.amount} for {self.object.contributor.name} on {self.object.date}'
        )
        
        return response


class FundUpdateView(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    """Update existing fund"""
    model = Fund
    template_name = 'funds/fund_form.html'
    fields = ['contributor', 'amount', 'date', 'description']
    success_url = reverse_lazy('fund-list')
    success_message = 'Fund updated successfully!'
    login_url = 'login'
    
    def get_queryset(self):
        return Fund.objects.filter(is_active=True)
    
    def form_valid(self, form):
        response = super().form_valid(form)
        
        # Log activity
        ActivityLog.objects.create(
            user=self.request.user,
            action='UPDATE',
            model_name='Fund',
            object_id=self.object.id,
            description=f'Updated fund to {self.object.amount} for {self.object.contributor.name} on {self.object.date}'
        )
        
        return response


class FundDeleteView(LoginRequiredMixin, SuccessMessageMixin, DeleteView):
    """Delete (soft delete) a fund"""
    model = Fund
    template_name = 'funds/fund_confirm_delete.html'
    success_url = reverse_lazy('fund-list')
    success_message = 'Fund deleted successfully!'
    login_url = 'login'
    
    def get_queryset(self):
        return Fund.objects.filter(is_active=True)
    
    def delete(self, request, *args, **kwargs):
        obj = self.get_object()
        
        # Log activity before soft delete
        ActivityLog.objects.create(
            user=self.request.user,
            action='DELETE',
            model_name='Fund',
            object_id=obj.id,
            description=f'Deleted fund of {obj.amount} for {obj.contributor.name}'
        )
        
        return super().delete(request, *args, **kwargs)
    
    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        
        # Log activity before deletion
        ActivityLog.objects.create(
            user=self.request.user,
            action='DELETE',
            model_name='Fund',
            object_id=self.object.id,
            description=f'Deleted fund of {self.object.amount} for {self.object.get_partner_display()}'
        )
        
        self.object.is_active = False
        self.object.save()
        return redirect(self.success_url)


# ==========================================
# EXPENSE VIEWS
# ==========================================

class ExpenseListView(LoginRequiredMixin, ListView):
    """List all expenses with filtering"""
    model = Expense
    template_name = 'expenses/expense_list.html'
    context_object_name = 'expenses'
    paginate_by = 20
    login_url = 'login'
    
    def get_queryset(self):
        queryset = Expense.objects.filter(is_active=True).order_by('-date')
        
        # Filter by category
        category = self.request.GET.get('category')
        if category:
            queryset = queryset.filter(category_id=category)
        
        # Filter by date range
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        
        if date_from:
            queryset = queryset.filter(date__gte=date_from)
        if date_to:
            queryset = queryset.filter(date__lte=date_to)
        
        # Search description
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(description__icontains=search) |
                Q(category__name__icontains=search)
            )
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        expenses = self.get_queryset()
        context['total_expenses'] = expenses.aggregate(
            total=Coalesce(Sum('amount'), 0, output_field=DecimalField())
        )['total']
        context['expenses_count'] = expenses.count()
        context['average_expense'] = expenses.aggregate(
            avg=Coalesce(Sum('amount') / Count('id'), 0, output_field=DecimalField())
        )['avg'] if expenses.exists() else 0
        context['categories'] = ExpenseCategory.objects.filter(is_active=True)
        
        return context


class ExpenseDetailView(LoginRequiredMixin, DetailView):
    """View expense details"""
    model = Expense
    template_name = 'expenses/expense_detail.html'
    context_object_name = 'expense'
    login_url = 'login'
    
    def get_queryset(self):
        return Expense.objects.filter(is_active=True)


class ExpenseCreateView(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    """Create a new expense"""
    model = Expense
    template_name = 'expenses/expense_form.html'
    fields = ['category', 'amount', 'date', 'description']
    success_url = reverse_lazy('expense-list')
    success_message = 'Expense recorded successfully!'
    login_url = 'login'
    
    def form_valid(self, form):
        form.instance.paid_by = self.request.user
        response = super().form_valid(form)
        
        # Handle document uploads
        self.handle_documents()
        
        # Log activity
        ActivityLog.objects.create(
            user=self.request.user,
            action='CREATE',
            model_name='Expense',
            object_id=self.object.id,
            description=f'Created expense of {self.object.amount} for {self.object.category.name} on {self.object.date}'
        )
        
        return response
    
    def handle_documents(self):
        """Handle multiple document uploads"""
        files = self.request.FILES.getlist('documents')
        for file in files:
            ExpenseDocument.objects.create(
                expense=self.object,
                file=file,
                uploaded_by=self.request.user
            )


class ExpenseUpdateView(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    """Update existing expense"""
    model = Expense
    template_name = 'expenses/expense_form.html'
    fields = ['category', 'amount', 'date', 'description']
    success_url = reverse_lazy('expense-list')
    success_message = 'Expense updated successfully!'
    login_url = 'login'
    
    def get_queryset(self):
        return Expense.objects.filter(is_active=True)
    
    def form_valid(self, form):
        response = super().form_valid(form)
        self.handle_documents()
        
        # Log activity
        ActivityLog.objects.create(
            user=self.request.user,
            action='UPDATE',
            model_name='Expense',
            object_id=self.object.id,
            description=f'Updated expense to {self.object.amount} for {self.object.category.name} on {self.object.date}'
        )
        
        return response
    
    def handle_documents(self):
        """Handle document uploads during update"""
        files = self.request.FILES.getlist('documents')
        for file in files:
            ExpenseDocument.objects.create(
                expense=self.object,
                file=file,
                uploaded_by=self.request.user
            )


class ExpenseDeleteView(LoginRequiredMixin, SuccessMessageMixin, DeleteView):
    """Delete (soft delete) an expense"""
    model = Expense
    template_name = 'expenses/expense_confirm_delete.html'
    success_url = reverse_lazy('expense-list')
    success_message = 'Expense deleted successfully!'
    login_url = 'login'
    
    def get_queryset(self):
        return Expense.objects.filter(is_active=True)
    
    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        
        # Log activity before deletion
        ActivityLog.objects.create(
            user=self.request.user,
            action='DELETE',
            model_name='Expense',
            object_id=self.object.id,
            description=f'Deleted expense of {self.object.amount} for {self.object.category.name}'
        )
        
        self.object.is_active = False
        self.object.save()
        return redirect(self.success_url)


# ==========================================
# CONTRIBUTOR VIEWS
# ==========================================

class ContributorListView(LoginRequiredMixin, ListView):
    """List all contributors"""
    model = Contributor
    template_name = 'contributors/contributor_list.html'
    context_object_name = 'contributors'
    login_url = 'login'
    
    def get_queryset(self):
        return Contributor.objects.all().order_by('name')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        contributors = self.get_queryset()
        
        context['active_count'] = contributors.filter(is_active=True).count()
        context['inactive_count'] = contributors.filter(is_active=False).count()
        
        # Add stats to each contributor
        for contributor in context['contributors']:
            funds = contributor.funds.filter(is_active=True)
            contributor.total_funds = funds.aggregate(
                total=Coalesce(Sum('amount'), 0, output_field=DecimalField())
            )['total']
            contributor.fund_count = funds.count()
        
        context['total_funds'] = Fund.objects.filter(is_active=True).aggregate(
            total=Coalesce(Sum('amount'), 0, output_field=DecimalField())
        )['total']
        
        if contributors.exists():
            context['avg_per_contributor'] = context['total_funds'] / contributors.count()
        else:
            context['avg_per_contributor'] = 0
        
        return context


class ContributorCreateView(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    """Create a new contributor"""
    model = Contributor
    template_name = 'contributors/contributor_form.html'
    fields = ['name', 'description']
    success_url = reverse_lazy('contributor-list')
    success_message = 'Contributor created successfully!'
    login_url = 'login'
    
    def form_valid(self, form):
        response = super().form_valid(form)
        
        # Log activity
        ActivityLog.objects.create(
            user=self.request.user,
            action='CREATE',
            model_name='Contributor',
            object_id=self.object.id,
            description=f'Created contributor: {self.object.name}'
        )
        
        return response


class ContributorUpdateView(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    """Update existing contributor"""
    model = Contributor
    template_name = 'contributors/contributor_form.html'
    fields = ['name', 'description']
    success_url = reverse_lazy('contributor-list')
    success_message = 'Contributor updated successfully!'
    login_url = 'login'
    
    def get_queryset(self):
        return Contributor.objects.all()
    
    def form_valid(self, form):
        response = super().form_valid(form)
        
        # Log activity
        ActivityLog.objects.create(
            user=self.request.user,
            action='UPDATE',
            model_name='Contributor',
            object_id=self.object.id,
            description=f'Updated contributor: {self.object.name}'
        )
        
        return response


class ContributorDeleteView(LoginRequiredMixin, SuccessMessageMixin, DeleteView):
    """Delete (soft delete) a contributor"""
    model = Contributor
    template_name = 'contributors/contributor_confirm_delete.html'
    success_url = reverse_lazy('contributor-list')
    success_message = 'Contributor deleted successfully!'
    login_url = 'login'
    
    def get_queryset(self):
        return Contributor.objects.all()
    
    def delete(self, request, *args, **kwargs):
        obj = self.get_object()
        
        # Check if contributor has active funds
        if obj.funds.filter(is_active=True).exists():
            messages.error(
                request,
                f'Cannot delete {obj.name}. There are funds associated with this contributor. Delete or reassign the funds first.'
            )
            return redirect('contributor-list')
        
        # Log activity
        ActivityLog.objects.create(
            user=request.user,
            action='DELETE',
            model_name='Contributor',
            object_id=obj.id,
            description=f'Deleted contributor: {obj.name}'
        )
        
        return super().delete(request, *args, **kwargs)


# ==========================================
# CATEGORY VIEWS
# ==========================================

class CategoryListView(LoginRequiredMixin, ListView):
    """List all expense categories"""
    model = ExpenseCategory
    template_name = 'categories/category_list.html'
    context_object_name = 'categories'
    login_url = 'login'
    
    def get_queryset(self):
        return ExpenseCategory.objects.all().order_by('name')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        categories = self.get_queryset()
        
        context['active_count'] = categories.filter(is_active=True).count()
        context['inactive_count'] = categories.filter(is_active=False).count()
        
        # Add stats to each category
        for category in context['categories']:
            expenses = category.expenses.filter(is_active=True)
            category.total_expenses = expenses.aggregate(
                total=Coalesce(Sum('amount'), 0, output_field=DecimalField())
            )['total']
            category.expense_count = expenses.count()
        
        context['total_expenses'] = Expense.objects.filter(is_active=True).aggregate(
            total=Coalesce(Sum('amount'), 0, output_field=DecimalField())
        )['total']
        
        if categories.exists():
            context['avg_per_category'] = context['total_expenses'] / categories.count()
        else:
            context['avg_per_category'] = 0
        
        return context


class CategoryCreateView(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    """Create a new category"""
    model = ExpenseCategory
    template_name = 'categories/category_form.html'
    fields = ['name', 'description']
    success_url = reverse_lazy('category-list')
    success_message = 'Category created successfully!'
    login_url = 'login'
    
    def form_valid(self, form):
        response = super().form_valid(form)
        
        # Log activity
        ActivityLog.objects.create(
            user=self.request.user,
            action='CREATE',
            model_name='ExpenseCategory',
            object_id=self.object.id,
            description=f'Created category: {self.object.name}'
        )
        
        return response


class CategoryUpdateView(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    """Update existing category"""
    model = ExpenseCategory
    template_name = 'categories/category_form.html'
    fields = ['name', 'description']
    success_url = reverse_lazy('category-list')
    success_message = 'Category updated successfully!'
    login_url = 'login'
    
    def form_valid(self, form):
        response = super().form_valid(form)
        
        # Log activity
        ActivityLog.objects.create(
            user=self.request.user,
            action='UPDATE',
            model_name='ExpenseCategory',
            object_id=self.object.id,
            description=f'Updated category: {self.object.name}'
        )
        
        return response


class CategoryDeleteView(LoginRequiredMixin, SuccessMessageMixin, DeleteView):
    """Delete (soft delete) a category"""
    model = ExpenseCategory
    template_name = 'categories/category_confirm_delete.html'
    success_url = reverse_lazy('category-list')
    success_message = 'Category deleted successfully!'
    login_url = 'login'
    
    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        
        # Log activity before deletion
        ActivityLog.objects.create(
            user=self.request.user,
            action='DELETE',
            model_name='ExpenseCategory',
            object_id=self.object.id,
            description=f'Deleted category: {self.object.name}'
        )
        
        self.object.is_active = False
        self.object.save()
        return redirect(self.success_url)


# ==========================================
# REPORT VIEWS
# ==========================================

class ReportDashboardView(LoginRequiredMixin, TemplateView):
    """Financial reports dashboard"""
    template_name = 'reports/report_dashboard.html'
    login_url = 'login'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get filters
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        category_id = self.request.GET.get('category')
        
        # Base querysets
        funds_qs = Fund.objects.filter(is_active=True)
        expenses_qs = Expense.objects.filter(is_active=True)
        
        # Apply filters
        if date_from:
            funds_qs = funds_qs.filter(date__gte=date_from)
            expenses_qs = expenses_qs.filter(date__gte=date_from)
        
        if date_to:
            funds_qs = funds_qs.filter(date__lte=date_to)
            expenses_qs = expenses_qs.filter(date__lte=date_to)
        
        if category_id:
            expenses_qs = expenses_qs.filter(category_id=category_id)
        
        # Summary
        context['total_funds'] = funds_qs.aggregate(
            total=Coalesce(Sum('amount'), 0, output_field=DecimalField())
        )['total']
        context['total_expenses'] = expenses_qs.aggregate(
            total=Coalesce(Sum('amount'), 0, output_field=DecimalField())
        )['total']
        context['net_balance'] = context['total_funds'] - context['total_expenses']
        
        # Monthly breakdown
        monthly_stats = []
        for expense in expenses_qs.order_by('date'):
            month_key = expense.date.replace(day=1)
            
            month_funds = funds_qs.filter(date__month=month_key.month, date__year=month_key.year).aggregate(
                total=Coalesce(Sum('amount'), 0, output_field=DecimalField())
            )['total']
            
            month_expenses = expenses_qs.filter(date__month=month_key.month, date__year=month_key.year).aggregate(
                total=Coalesce(Sum('amount'), 0, output_field=DecimalField())
            )['total']
            
            if not any(m['month'] == month_key for m in monthly_stats):
                percent_spent = (month_expenses / month_funds * 100) if month_funds > 0 else 0
                monthly_stats.append({
                    'month': month_key,
                    'income': month_funds,
                    'expenses': month_expenses,
                    'balance': month_funds - month_expenses,
                    'percent_spent': percent_spent,
                })
        
        context['monthly_stats'] = monthly_stats
        
        # Category breakdown
        category_stats = []
        for category in ExpenseCategory.objects.filter(is_active=True):
            cat_expenses = expenses_qs.filter(category=category)
            cat_total = cat_expenses.aggregate(
                total=Coalesce(Sum('amount'), 0, output_field=DecimalField())
            )['total']
            
            if cat_total > 0:
                percent = (cat_total / context['total_expenses'] * 100) if context['total_expenses'] > 0 else 0
                category_stats.append({
                    'category_name': category.name,
                    'total': cat_total,
                    'count': cat_expenses.count(),
                    'average': cat_total / cat_expenses.count() if cat_expenses.exists() else 0,
                    'percent': percent,
                })
        
        context['category_stats'] = category_stats
        context['categories'] = ExpenseCategory.objects.filter(is_active=True)
        
        return context


@login_required(login_url='login')
def report_export(request):
    """Export reports as CSV or Excel"""
    format_type = request.GET.get('format', 'csv')
    
    # Get filters and data
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    category_id = request.GET.get('category')
    
    expenses_qs = Expense.objects.filter(is_active=True).order_by('-date')
    
    if date_from:
        expenses_qs = expenses_qs.filter(date__gte=date_from)
    if date_to:
        expenses_qs = expenses_qs.filter(date__lte=date_to)
    if category_id:
        expenses_qs = expenses_qs.filter(category_id=category_id)
    
    if format_type == 'excel':
        return export_excel(expenses_qs)
    else:
        return export_csv(expenses_qs)


def export_csv(expenses_qs):
    """Export expenses as CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="expense_report.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['ID', 'Date', 'Category', 'Amount', 'Paid By', 'Description', 'Documents'])
    
    for expense in expenses_qs:
        writer.writerow([
            expense.id,
            expense.date,
            expense.category.name,
            expense.amount,
            expense.paid_by.get_full_name() or expense.paid_by.username,
            expense.description,
            expense.documents.count(),
        ])
    
    return response


def export_excel(expenses_qs):
    """Export expenses as Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expenses"
    
    # Headers
    headers = ['ID', 'Date', 'Category', 'Amount', 'Paid By', 'Description', 'Documents']
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Data
    for expense in expenses_qs:
        ws.append([
            expense.id,
            expense.date,
            expense.category.name,
            float(expense.amount),
            expense.paid_by.get_full_name() or expense.paid_by.username,
            expense.description,
            expense.documents.count(),
        ])
    
    # Auto-adjust columns
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 10
    
    # Save to BytesIO
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="expense_report.xlsx"'
    
    wb.save(response)
    return response


# ==========================================
# ACTIVITY LOG VIEWS
# ==========================================

class ActivityLogListView(LoginRequiredMixin, ListView):
    """View activity logs"""
    model = ActivityLog
    template_name = 'activity_logs/activity_log_list.html'
    context_object_name = 'activity_logs'
    paginate_by = 50
    login_url = 'login'
    
    def get_queryset(self):
        queryset = ActivityLog.objects.all().order_by('-timestamp')
        
        # Filter by action
        action = self.request.GET.get('action')
        if action:
            queryset = queryset.filter(action=action)
        
        # Filter by user
        user_id = self.request.GET.get('user')
        if user_id:
            queryset = queryset.filter(user_id=user_id)
        
        # Filter by model
        model_name = self.request.GET.get('model_name')
        if model_name:
            queryset = queryset.filter(model_name=model_name)
        
        # Filter by date range
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        
        if date_from:
            queryset = queryset.filter(timestamp__date__gte=date_from)
        if date_to:
            queryset = queryset.filter(timestamp__date__lte=date_to)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['actions'] = ActivityLog.ACTION_CHOICES
        context['models'] = ActivityLog.objects.values_list('model_name', flat=True).distinct()
        context['users'] = User.objects.filter(
            activitylog__user__isnull=False
        ).distinct()
        return context


# ==========================================
# DELETED RECORDS VIEWS (SUPERADMIN ONLY)
# ==========================================

class DeletedRecordsListView(LoginRequiredMixin, ListView):
    """View deleted records (soft deleted) - Superadmin only"""
    template_name = 'deleted_records/deleted_records_list.html'
    context_object_name = 'deleted_records'
    paginate_by = 50
    login_url = 'login'
    
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_superuser:
            from django.core.exceptions import PermissionDenied
            raise PermissionDenied("Only superadmin can access deleted records.")
        return super().dispatch(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get counts of deleted records by model
        deleted_funds = Fund.objects.filter(is_active=False).count()
        deleted_expenses = Expense.objects.filter(is_active=False).count()
        deleted_categories = ExpenseCategory.objects.filter(is_active=False).count()
        
        context['deleted_funds_count'] = deleted_funds
        context['deleted_expenses_count'] = deleted_expenses
        context['deleted_categories_count'] = deleted_categories
        context['total_deleted'] = deleted_funds + deleted_expenses + deleted_categories
        
        return context


class DeletedFundsView(LoginRequiredMixin, ListView):
    """View deleted funds - Superadmin only"""
    model = Fund
    template_name = 'deleted_records/deleted_funds.html'
    context_object_name = 'funds'
    paginate_by = 20
    login_url = 'login'
    
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_superuser:
            from django.core.exceptions import PermissionDenied
            raise PermissionDenied("Only superadmin can access deleted records.")
        return super().dispatch(request, *args, **kwargs)
    
    def get_queryset(self):
        return Fund.objects.filter(is_active=False).order_by('-updated_at')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        funds = self.get_queryset()
        context['total_amount'] = funds.aggregate(
            total=Coalesce(Sum('amount'), 0, output_field=DecimalField())
        )['total']
        context['funds_count'] = funds.count()
        return context


class DeletedExpensesView(LoginRequiredMixin, ListView):
    """View deleted expenses - Superadmin only"""
    model = Expense
    template_name = 'deleted_records/deleted_expenses.html'
    context_object_name = 'expenses'
    paginate_by = 20
    login_url = 'login'
    
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_superuser:
            from django.core.exceptions import PermissionDenied
            raise PermissionDenied("Only superadmin can access deleted records.")
        return super().dispatch(request, *args, **kwargs)
    
    def get_queryset(self):
        return Expense.objects.filter(is_active=False).order_by('-updated_at')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        expenses = self.get_queryset()
        context['total_amount'] = expenses.aggregate(
            total=Coalesce(Sum('amount'), 0, output_field=DecimalField())
        )['total']
        context['expenses_count'] = expenses.count()
        return context


class DeletedCategoriesView(LoginRequiredMixin, ListView):
    """View deleted categories - Superadmin only"""
    model = ExpenseCategory
    template_name = 'deleted_records/deleted_categories.html'
    context_object_name = 'categories'
    paginate_by = 20
    login_url = 'login'
    
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_superuser:
            from django.core.exceptions import PermissionDenied
            raise PermissionDenied("Only superadmin can access deleted records.")
        return super().dispatch(request, *args, **kwargs)
    
    def get_queryset(self):
        return ExpenseCategory.objects.filter(is_active=False).order_by('-updated_at')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        categories = self.get_queryset()
        context['categories_count'] = categories.count()
        return context


class RestoreRecordView(LoginRequiredMixin, View):
    """Restore a deleted record - Superadmin only"""
    login_url = 'login'
    
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_superuser:
            from django.core.exceptions import PermissionDenied
            raise PermissionDenied("Only superadmin can restore deleted records.")
        return super().dispatch(request, *args, **kwargs)
    
    def post(self, request, *args, **kwargs):
        model_name = request.POST.get('model_name')
        object_id = request.POST.get('object_id')
        
        try:
            if model_name == 'Fund':
                obj = Fund.objects.get(id=object_id)
                obj.is_active = True
                obj.save()
                ActivityLog.objects.create(
                    user=request.user,
                    action='UPDATE',
                    model_name='Fund',
                    object_id=obj.id,
                    description=f'Restored deleted fund of {obj.amount} for {obj.contributor.name}'
                )
                messages.success(request, f'Fund #{object_id} has been restored.')
                return redirect('deleted-funds')
            
            elif model_name == 'Expense':
                obj = Expense.objects.get(id=object_id)
                obj.is_active = True
                obj.save()
                ActivityLog.objects.create(
                    user=request.user,
                    action='UPDATE',
                    model_name='Expense',
                    object_id=obj.id,
                    description=f'Restored deleted expense of {obj.amount} for {obj.category.name}'
                )
                messages.success(request, f'Expense #{object_id} has been restored.')
                return redirect('deleted-expenses')
            
            elif model_name == 'ExpenseCategory':
                obj = ExpenseCategory.objects.get(id=object_id)
                obj.is_active = True
                obj.save()
                ActivityLog.objects.create(
                    user=request.user,
                    action='UPDATE',
                    model_name='ExpenseCategory',
                    object_id=obj.id,
                    description=f'Restored deleted category: {obj.name}'
                )
                messages.success(request, f'Category #{object_id} has been restored.')
                return redirect('deleted-categories')
        
        except (Fund.DoesNotExist, Expense.DoesNotExist, ExpenseCategory.DoesNotExist):
            messages.error(request, 'Record not found.')
        
        return redirect('deleted-records')


# ==========================================
# TODO (PLANNED EXPENSE) VIEWS
# ==========================================

class TodoListView(LoginRequiredMixin, ListView):
    """List all todos with filtering"""
    model = Todo
    template_name = 'todos/todo_list.html'
    context_object_name = 'todos'
    paginate_by = 20
    login_url = 'login'
    
    def get_queryset(self):
        queryset = Todo.objects.filter(is_active=True).order_by('-target_date')
        
        # Filter by status
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)
        
        # Filter by category
        category = self.request.GET.get('category')
        if category:
            queryset = queryset.filter(category_id=category)
        
        # Filter by date range
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        
        if date_from:
            queryset = queryset.filter(target_date__gte=date_from)
        if date_to:
            queryset = queryset.filter(target_date__lte=date_to)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        todos = self.get_queryset()
        
        # Calculate stats
        context['pending_count'] = todos.filter(status='PENDING').count()
        context['completed_count'] = todos.filter(status='COMPLETED').count()
        context['cancelled_count'] = todos.filter(status='CANCELLED').count()
        context['estimated_total'] = todos.filter(
            estimated_amount__isnull=False
        ).aggregate(
            total=Coalesce(Sum('estimated_amount'), 0, output_field=DecimalField())
        )['total']
        context['overdue_count'] = sum(
            1 for todo in todos.filter(status='PENDING') if todo.is_overdue()
        )
        context['categories'] = ExpenseCategory.objects.filter(is_active=True)
        
        return context


class TodoCreateView(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    """Create a new todo"""
    model = Todo
    template_name = 'expenses/todo_form.html'
    fields = ['title', 'description', 'category', 'estimated_amount', 'target_date']
    success_url = reverse_lazy('todo-list')
    success_message = 'Todo created successfully!'
    login_url = 'login'
    
    def form_valid(self, form):
        form.instance.created_by = self.request.user
        response = super().form_valid(form)
        
        # Log activity
        ActivityLog.objects.create(
            user=self.request.user,
            action='CREATE',
            model_name='Todo',
            object_id=self.object.id,
            description=f'Created todo: {self.object.title} (Target: {self.object.target_date})'
        )
        
        return response


class TodoUpdateView(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    """Update existing todo"""
    model = Todo
    template_name = 'expenses/todo_form.html'
    fields = ['title', 'description', 'category', 'estimated_amount', 'target_date', 'status']
    success_url = reverse_lazy('todo-list')
    success_message = 'Todo updated successfully!'
    login_url = 'login'
    
    def get_queryset(self):
        return Todo.objects.filter(is_active=True)
    
    def form_valid(self, form):
        response = super().form_valid(form)
        
        # Log activity
        ActivityLog.objects.create(
            user=self.request.user,
            action='UPDATE',
            model_name='Todo',
            object_id=self.object.id,
            description=f'Updated todo: {self.object.title} (Status: {self.object.get_status_display()})'
        )
        
        return response


class TodoDeleteView(LoginRequiredMixin, SuccessMessageMixin, DeleteView):
    """Soft delete a todo"""
    model = Todo
    template_name = 'expenses/todo_confirm_delete.html'
    success_url = reverse_lazy('todo-list')
    success_message = 'Todo deleted successfully!'
    login_url = 'login'
    
    def get_queryset(self):
        return Todo.objects.filter(is_active=True)
    
    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        
        # Log activity
        ActivityLog.objects.create(
            user=self.request.user,
            action='DELETE',
            model_name='Todo',
            object_id=self.object.id,
            description=f'Deleted todo: {self.object.title}'
        )
        
        self.object.is_active = False
        self.object.save()
        return redirect(self.success_url)


class TodoConvertView(LoginRequiredMixin, View):
    """Convert todo to expense"""
    login_url = 'login'
    
    def get(self, request, pk):
        todo = get_object_or_404(Todo, id=pk, is_active=True)
        context = {
            'todo': todo,
        }
        return render(request, 'expenses/todo_confirm_convert.html', context)
    
    def post(self, request, pk):
        todo = get_object_or_404(Todo, id=pk, is_active=True)
        
        # Create expense from todo
        expense = Expense.objects.create(
            category=todo.category,
            amount=todo.estimated_amount or 0,
            date=timezone.now().date(),
            description=todo.description or f"From todo: {todo.title}",
            paid_by=request.user
        )
        
        # Mark todo as completed
        todo.status = 'COMPLETED'
        todo.save()
        
        # Log activities
        ActivityLog.objects.create(
            user=request.user,
            action='CREATE',
            model_name='Expense',
            object_id=expense.id,
            description=f'Created expense from todo: {todo.title}'
        )
        
        ActivityLog.objects.create(
            user=request.user,
            action='UPDATE',
            model_name='Todo',
            object_id=todo.id,
            description=f'Converted todo to expense: {todo.title}'
        )
        
        messages.success(request, f'Todo converted to expense successfully!')
        return redirect('expense-detail', pk=expense.id)


# -----------------------
# DOCUMENT MANAGEMENT VIEWS
# -----------------------

class ExpenseDocumentDeleteView(LoginRequiredMixin, SuccessMessageMixin, DeleteView):
    """Delete an expense document"""
    model = ExpenseDocument
    template_name = 'expenses/expense_document_confirm_delete.html'
    success_message = 'Document deleted successfully!'
    login_url = 'login'
    
    def get_success_url(self):
        return reverse_lazy('expense-detail', kwargs={'pk': self.object.expense.id})
    
    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        expense_id = self.object.expense.id
        
        # Log activity
        ActivityLog.objects.create(
            user=request.user,
            action='DELETE',
            model_name='ExpenseDocument',
            object_id=self.object.id,
            description=f'Deleted document from expense #{expense_id}'
        )
        
        response = super().delete(request, *args, **kwargs)
        return response


@login_required
def add_expense_documents(request, pk):
    """Add multiple documents to an existing expense"""
    expense = get_object_or_404(Expense, pk=pk, is_active=True)
    
    if request.method == 'POST':
        files = request.FILES.getlist('documents')
        
        if not files:
            messages.error(request, 'Please select at least one file.')
            return redirect('expense-detail', pk=pk)
        
        for file in files:
            ExpenseDocument.objects.create(
                expense=expense,
                file=file,
                uploaded_by=request.user
            )
            
            # Log activity
            ActivityLog.objects.create(
                user=request.user,
                action='CREATE',
                model_name='ExpenseDocument',
                description=f'Added document to expense #{expense.id}'
            )
        
        messages.success(request, f'Successfully added {len(files)} document(s) to expense!')
        return redirect('expense-detail', pk=pk)
    
    return redirect('expense-detail', pk=pk)

@login_required
def download_expense_document(request, pk):
    """Download an expense document with proper MIME type"""
    doc = get_object_or_404(ExpenseDocument, pk=pk)
    
    try:
        file_path = doc.file.path
        file_name = doc.file.name.split('/')[-1]
        
        # Determine MIME type based on file extension
        import mimetypes
        mime_type, _ = mimetypes.guess_type(file_path)
        if not mime_type:
            mime_type = 'application/octet-stream'
        
        with open(file_path, 'rb') as f:
            response = HttpResponse(f.read(), content_type=mime_type)
            response['Content-Disposition'] = f'attachment; filename="{file_name}"'
            return response
    except FileNotFoundError:
        messages.error(request, 'File not found.')
        return redirect('expense-detail', pk=doc.expense.id)


@login_required
def view_expense_document(request, pk):
    """View an expense document (inline) - opens in new tab with correct MIME type"""
    doc = get_object_or_404(ExpenseDocument, pk=pk)
    
    try:
        file_path = doc.file.path
        file_name = doc.file.name.split('/')[-1]
        
        # Determine MIME type based on file extension
        import mimetypes
        mime_type, _ = mimetypes.guess_type(file_path)
        if not mime_type:
            mime_type = 'application/octet-stream'
        
        with open(file_path, 'rb') as f:
            response = HttpResponse(f.read(), content_type=mime_type)
            # Use inline to display in browser instead of download
            response['Content-Disposition'] = f'inline; filename="{file_name}"'
            return response
    except FileNotFoundError:
        messages.error(request, 'File not found.')
        return redirect('expense-detail', pk=doc.expense.id)